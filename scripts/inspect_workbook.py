import openpyxl, json
from pathlib import Path
path = Path('xlsx/CRA_Maturity_Model_TOOL.xlsx')
wb = openpyxl.load_workbook(path, data_only=True)
print('SHEETS', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n=== {name} ===')
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True):
        print(row)
